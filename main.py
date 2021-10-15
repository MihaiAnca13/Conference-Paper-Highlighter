import os
from selenium import webdriver
from time import sleep
from selenium.common.exceptions import NoSuchElementException
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.colors import COLOR_INDEX

RED = COLOR_INDEX[2]
BLUE = COLOR_INDEX[4]
GREEN = COLOR_INDEX[3]
YELLOW = COLOR_INDEX[5]
colors = [RED, BLUE, GREEN, YELLOW]


def removeFormatting(ws):
    for key, _ in ws.conditional_formatting._cf_rules.items():
        del ws.conditional_formatting[key.sqref]

    # ws is not the worksheet name, but the worksheet object
    for row in ws.iter_rows():
        for cell in row:
            cell.style = 'Normal'


if not os.path.isfile('sample.xlsx'):
    wb = Workbook()
    wb.save('sample.xlsx')
else:
    wb = load_workbook('sample.xlsx')

# grab the active worksheet
ws = wb.active

removeFormatting(ws)

ans = input('Download list? [y/n]')
if ans == 'y':
    URL = "https://openreview.net/"

    browser = webdriver.Chrome()
    browser.get(URL)

    # wait for page to load
    sleep(2)

    # select conference
    confs = browser.find_element_by_id('all-venues')
    confs = confs.find_elements_by_tag_name('a')
    for i, c in enumerate(confs):
        print(i, c.text, c.get_attribute('href'))

    id = int(input())
    new_link = confs[id].get_attribute('href')

    browser.get(new_link)

    # navigate the pages
    while True:
        try:
            options = browser.find_element_by_class_name('venues-list').find_elements_by_tag_name('a')
        except NoSuchElementException:
            break
        for i, o in enumerate(options):
            print(i, o.text, o.get_attribute('href'))

        id = int(input())
        new_link = options[id].get_attribute('href')

        browser.get(new_link)

    # grab all the papers!
    id = browser.current_url
    id = id[id.find('id=')+3:]
    links = []
    links.append(f"https://api.openreview.net/notes?invitation={id}/.*/Submission&details=replyCount,invitation,original&offset=0&includeCount=false&limit=5000")
    # links.append(f"https://api.openreview.net/notes?invitation={id}/Paper.*/-/Decision&details=replyCount,invitation,original&offset=0&includeCount=false&limit=5000")
    links.append(f"https://api.openreview.net/notes?invitation={id}/-/Blind_Submission&details=replyCount,invitation,original&offset=0&includeCount=false&limit=5000")

    for link in links:
        resp = requests.get(link).json()
        if 'count' in resp:
            for i, note in enumerate(resp['notes']):
                if 'decision' in note['content'] and note['content']['decision'] == 'Reject':
                    continue
                title = note['content']['title']
                authors = note['content']['authors']
                keywords = note['content']['keywords']

                print(title, authors, keywords)
                ws[f'A{i+1}'] = title
                ws[f'B{i+1}'] = ", ".join(authors)
                ws[f'C{i+1}'] = ", ".join(keywords)

    browser.close()
# Save the file
wb.save("sample.xlsx")

i = 0
ws['I1'] = "Text to highlight (limited to the first 5000 rows):"
for j in range(2, 6):
    ws[f'I{j}'] = ""
while True:
    print(f'Text to highlight{i+1}/4 (empty for skip): ')
    text_to_find = input()
    if text_to_find == '':
        break

    fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type='solid')
    dxf = DifferentialStyle(fill=fill)
    ws[f'I{2+i}'] = text_to_find
    ws[f'I{2 + i}'].fill = fill
    rule = Rule(type='containsText', operator='containsText', text=text_to_find, stopIfTrue=False)
    rule.dxf = dxf
    rule.formula = [f"NOT(ISERROR(SEARCH(\"{text_to_find}\",A1)))"]

    ws.conditional_formatting.add('A1:D5000', rule)

    i += 1
    if i == 4:
        break

# Save the file
wb.save("sample.xlsx")
wb.close()