from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.colors import COLOR_INDEX

RED = COLOR_INDEX[2]
BLUE = COLOR_INDEX[4]
GREEN = COLOR_INDEX[3]
YELLOW = COLOR_INDEX[5]
colors = [RED, BLUE, GREEN, YELLOW]


def removeFormatting(ws):
    for key, _ in ws.conditional_formatting._cf_rules.items():
        del ws.conditional_formatting[key.sqref]

    # ws is not the worksheet name, but the worksheet object
    for row in ws.iter_rows():
        for cell in row:
            cell.style = 'Normal'


wb = load_workbook('sample.xlsx')

# grab the active worksheet
ws = wb.active

removeFormatting(ws)

redFill = PatternFill(start_color=colors[3], end_color=colors[3], fill_type='solid')
dxf = DifferentialStyle(fill=redFill)
text_to_find = '2'
ws['I1'] = text_to_find
rule = Rule(type='containsText', operator='containsText', text=text_to_find, stopIfTrue=False)
rule.dxf = dxf

ws.conditional_formatting.add('A1:F5000', rule)

# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
#
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
